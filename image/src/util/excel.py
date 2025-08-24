from openpyxl.styles import PatternFill, Font
import pandas as pd
from io import BytesIO

from src.util.format import format

def create_file(content=list[tuple[str, pd.DataFrame]], path: str='') -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        for sheetname, df in content:
            create_sheet(writer, sheetname, df)

    output.seek(0)
    binary = output.read()

    if path:
        with open(path, 'wb') as f:
            f.write(binary)
    
    return binary

def create_sheet(writer: pd.ExcelWriter, sheetname: str, df: pd.DataFrame):
    df.reset_index(drop=True, inplace=True)
    df.to_excel(writer, sheet_name=sheetname, index=False)
            
    worksheet = writer.sheets[sheetname]
    
    header_pattern, header_font = get_style(format['header'])
    even_pattern, even_font = get_style(format['body']['even'])
    odd_pattern, odd_font = get_style(format['body']['odd'])
    
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(1, col)
        cell.fill = header_pattern
        cell.font = header_font

    for row in range(2, worksheet.max_row + 1):
        pattern = even_pattern if row % 2 == 0 else odd_pattern
        font = even_font if row % 2 == 0 else odd_font
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row, col)
            cell.fill = pattern
            cell.font = font

def get_style(row_format):
    patten = PatternFill(start_color=row_format['bg_color'], end_color=row_format['bg_color'], fill_type='solid')
    font = Font(color=row_format['font_color'], bold=False)
    return patten, font